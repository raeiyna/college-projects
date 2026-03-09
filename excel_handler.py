import os
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import mysql.connector

def get_excel_path(user_id):
    # Save the file to a specific location on the Desktop
    desktop_path = os.path.join(os.path.expanduser("~"), "Desktop", "AXIS EXCELS", user_id)
    if not os.path.exists(desktop_path):
        os.makedirs(desktop_path)
    return os.path.join(desktop_path, f"student_performance_{user_id}.xlsx")

def excel_exists(user_id):
    return os.path.exists(get_excel_path(user_id))

def load_excel_path(user_id):
    return get_excel_path(user_id)

def is_valid_score(score):
    """Check if a score is valid (not empty, not None, and can be converted to float)"""
    if score is None:
        return False
    
    score_str = str(score).strip()
    if not score_str or score_str.lower() in ['', 'none', 'null', 'n/a', '-']:
        return False
    
    try:
        # Handle percentage format like "5/5 (100%)" or "3/5 (60%)"
        if '(' in score_str and ')' in score_str:
            # Extract the percentage value from parentheses
            start = score_str.find('(') + 1
            end = score_str.find('%')
            if start > 0 and end > start:
                percentage_str = score_str[start:end].strip()
                float(percentage_str)
                return True
        
        # Handle regular numeric scores
        float(score_str)
        return True
    except (ValueError, TypeError):
        return False
    

def extract_numeric_score(score):
    """Extract numeric value from score, handling percentage format"""
    if score is None:
        return None
    
    score_str = str(score).strip()
    if not score_str:
        return None
    
    try:
        # Handle percentage format like "5/5 (100%)" or "3/5 (60%)"
        if '(' in score_str and ')' in score_str:
            # Extract the percentage value from parentheses
            start = score_str.find('(') + 1
            end = score_str.find('%')
            if start > 0 and end > start:
                percentage_str = score_str[start:end].strip()
                return float(percentage_str)
        
        # Handle regular numeric scores
        return float(score_str)
    except (ValueError, TypeError):
        return None

def calculate_gwa_with_available_scores(prelim, midterm, finals):
    """
    Enhanced GWA calculation that handles missing scores intelligently.
    - Uses proportional weighting based on available scores
    - Returns None if no valid scores are available
    - Provides detailed calculation info for transparency
    """
    try:
        # Standard weights
        standard_weights = {
            'prelim': 0.30,    # 30%
            'midterm': 0.30,   # 30% 
            'finals': 0.40     # 40%
        }
        
        # Check which scores are valid and extract numeric values
        valid_scores = {}
        available_weights = {}
        
        if is_valid_score(prelim):
            numeric_score = extract_numeric_score(prelim)
            if numeric_score is not None:
                valid_scores['prelim'] = numeric_score
                available_weights['prelim'] = standard_weights['prelim']
            
        if is_valid_score(midterm):
            numeric_score = extract_numeric_score(midterm)
            if numeric_score is not None:
                valid_scores['midterm'] = numeric_score
                available_weights['midterm'] = standard_weights['midterm']
            
        if is_valid_score(finals):
            numeric_score = extract_numeric_score(finals)
            if numeric_score is not None:
                valid_scores['finals'] = numeric_score
                available_weights['finals'] = standard_weights['finals']
        
        # Return None if no valid scores (instead of 0)
        if not valid_scores:
            return None
        
        # Calculate proportional weights
        total_available_weight = sum(available_weights.values())
        proportional_weights = {
            exam: weight / total_available_weight 
            for exam, weight in available_weights.items()
        }
        
        # Calculate weighted GWA
        gwa = sum(
            valid_scores[exam] * proportional_weights[exam] 
            for exam in valid_scores.keys()
        )
        
        return round(gwa, 2)
        
    except Exception as e:
        print(f"Error calculating GWA: {e}")
        return None

def get_gwa_calculation_details(prelim, midterm, finals):
    """
    Provides detailed information about how GWA was calculated
    for transparency and debugging purposes
    """
    details = {
        'has_prelim': is_valid_score(prelim),
        'has_midterm': is_valid_score(midterm),
        'has_finals': is_valid_score(finals),
        'calculation_method': '',
        'weights_used': {}
    }
    
    valid_count = sum([details['has_prelim'], details['has_midterm'], details['has_finals']])
    
    if valid_count == 0:
        details['calculation_method'] = 'No valid scores available'
    elif valid_count == 3:
        details['calculation_method'] = 'All scores available (30%, 30%, 40%)'
        details['weights_used'] = {'prelim': '30%', 'midterm': '30%', 'finals': '40%'}
    elif valid_count == 2:
        if details['has_prelim'] and details['has_midterm']:
            details['calculation_method'] = 'Prelim + Midterm only (50%, 50%)'
            details['weights_used'] = {'prelim': '50%', 'midterm': '50%'}
        elif details['has_prelim'] and details['has_finals']:
            details['calculation_method'] = 'Prelim + Finals only (42.9%, 57.1%)'
            details['weights_used'] = {'prelim': '42.9%', 'finals': '57.1%'}
        elif details['has_midterm'] and details['has_finals']:
            details['calculation_method'] = 'Midterm + Finals only (42.9%, 57.1%)'
            details['weights_used'] = {'midterm': '42.9%', 'finals': '57.1%'}
    else:  # valid_count == 1
        if details['has_prelim']:
            details['calculation_method'] = 'Prelim only (100%)'
            details['weights_used'] = {'prelim': '100%'}
        elif details['has_midterm']:
            details['calculation_method'] = 'Midterm only (100%)'
            details['weights_used'] = {'midterm': '100%'}
        elif details['has_finals']:
            details['calculation_method'] = 'Finals only (100%)'
            details['weights_used'] = {'finals': '100%'}
    
    return details

def is_passing_grade(gwa, passing_threshold=75.0):
    """
    Enhanced passing grade determination
    - Returns None if GWA is None (no valid scores)
    - Otherwise returns True/False for pass/fail
    """
    if gwa is None:
        return None
    return gwa >= passing_threshold

def get_grade_status_text(gwa, passing_threshold=75.0):
    """Get appropriate status text based on GWA"""
    if gwa is None:
        return "INCOMPLETE"
    elif gwa >= passing_threshold:
        return "PASS"
    else:
        return "FAIL"

def get_all_courses_from_database(user_id):
    """Get ALL courses from both semesters for this faculty member"""
    try:
        connection = mysql.connector.connect(
            host="localhost",
            user="root",
            password="TUPCcoet@23!",
            database="axis"
        )
        cursor = connection.cursor(dictionary=True)
        
        # Get ALL courses for this faculty from both semesters
        cursor.execute("""
            SELECT DISTINCT subject_code, semester 
            FROM examination_form 
            WHERE created_by = %s
            ORDER BY semester, subject_code
        """, (user_id,))
        
        results = cursor.fetchall()
        
        # Organize courses by semester
        courses_by_semester = {"1st Semester": [], "2nd Semester": []}
        all_courses = set()
        
        for row in results:
            semester = row["semester"]
            course = row["subject_code"]
            courses_by_semester[semester].append(course)
            all_courses.add(course)
        
        cursor.close()
        connection.close()
        
        return courses_by_semester, list(all_courses)
        
    except mysql.connector.Error as err:
        print(f"Database Error in get_all_courses_from_database: {err}")
        return {"1st Semester": [], "2nd Semester": []}, []
    except Exception as e:
        print(f"Error in get_all_courses_from_database: {e}")
        return {"1st Semester": [], "2nd Semester": []}, []

def get_course_statistics(user_id, course, semester, populate_callback, table_widget):
    """Get enhanced statistics for a specific course including incomplete records"""
    try:
        # Use the populate callback to get data
        populate_callback(table_widget, course, semester)
        
        if table_widget.rowCount() == 0:
            return {
                "passing_students": 0, 
                "failing_students": 0, 
                "incomplete_students": 0,
                "total_students": 0
            }
        
        # Find exam columns
        prelim_col = midterm_col = finals_col = -1
        
        for col in range(table_widget.columnCount()):
            header = table_widget.horizontalHeaderItem(col)
            if header:
                header_text = header.text().upper()
                if "PRELIM" in header_text:
                    prelim_col = col
                elif "MIDTERM" in header_text:
                    midterm_col = col
                elif "FINAL" in header_text:
                    finals_col = col
        
        # Calculate enhanced statistics
        passing_students = 0
        failing_students = 0
        incomplete_students = 0
        
        for row in range(table_widget.rowCount()):
            prelim_score = ""
            midterm_score = ""
            finals_score = ""
            
            if prelim_col >= 0:
                item = table_widget.item(row, prelim_col)
                prelim_score = item.text().strip() if item else ""
                
            if midterm_col >= 0:
                item = table_widget.item(row, midterm_col)
                midterm_score = item.text().strip() if item else ""
                
            if finals_col >= 0:
                item = table_widget.item(row, finals_col)
                finals_score = item.text().strip() if item else ""
            
            # Calculate GWA for this student
            gwa = calculate_gwa_with_available_scores(prelim_score, midterm_score, finals_score)
            
            # Categorize student based on GWA result
            if gwa is None:
                incomplete_students += 1
            elif is_passing_grade(gwa):
                passing_students += 1
            else:
                failing_students += 1
        
        return {
            "passing_students": passing_students,
            "failing_students": failing_students,
            "incomplete_students": incomplete_students,
            "total_students": table_widget.rowCount()
        }
        
    except Exception as e:
        print(f"Error getting course statistics: {e}")
        return {
            "passing_students": 0, 
            "failing_students": 0, 
            "incomplete_students": 0,
            "total_students": 0
        }

def save_table_to_excel(table_widget, user_id, selected_course=None, selected_semester=None, course_sort_combobox=None, populate_callback=None):
    try:
        # Create a new workbook and remove the default sheet
        workbook = openpyxl.Workbook()
        workbook.remove(workbook.active)  # Remove default sheet

        # Enhanced styling for GWA features
        header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")  # Blue header
        header_font = Font(color="FFFFFF", bold=True, size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")
        cell_alignment = Alignment(horizontal="center", vertical="center")
        title_font = Font(bold=True, size=16, color="000000")
        subtitle_font = Font(bold=True, size=12, color="0066CC")
        
        # Enhanced GWA specific styling
        gwa_header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")  # Green for GWA
        pass_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")  # Light green for pass
        fail_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")  # Light red for fail
        incomplete_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")  # Light orange for incomplete
        pass_font = Font(color="2E7D32", bold=True)  # Dark green text
        fail_font = Font(color="C62828", bold=True)  # Dark red text
        incomplete_font = Font(color="E65100", bold=True)  # Orange text
        
        # Grading system styling
        grading_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")  # Light yellow
        grading_font = Font(color="E65100", bold=True, size=10)  # Orange text
        
        # Add borders for cleaner look
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Get ALL courses from database
        courses_by_semester, all_unique_courses = get_all_courses_from_database(user_id)
        
        print(f"DEBUG: 1st Semester courses: {courses_by_semester['1st Semester']}")
        print(f"DEBUG: 2nd Semester courses: {courses_by_semester['2nd Semester']}")
        print(f"DEBUG: All unique courses: {all_unique_courses}")
        
        # Create the single Summary sheet
        summary_sheet = workbook.create_sheet(title="Summary")
        
        # Add logo/title section
        summary_sheet.row_dimensions[1].height = 40
        summary_sheet.cell(row=1, column=1, value="AXIS - Enhanced Student Performance Dashboard with Smart GWA")
        summary_sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
        title_cell = summary_sheet.cell(row=1, column=1)
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        title_cell.fill = PatternFill(start_color="E5E5E5", end_color="E5E5E5", fill_type="solid")
        
        # Add enhanced instructions
        summary_sheet.cell(row=2, column=1, value="Smart GWA Calculation: Automatically adjusts weights based on available scores | Green: Pass, Red: Fail, Orange: Incomplete")
        summary_sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=7)
        summary_sheet.cell(row=2, column=1).alignment = Alignment(horizontal="center")
        summary_sheet.cell(row=2, column=1).font = Font(italic=True)
        
        # Store information about which sheets were actually created
        created_sheets = []
        course_statistics = {}
        
        # Create individual course sheets for each semester
        for semester in ["1st Semester", "2nd Semester"]:
            semester_prefix = semester[:3]  # "1st" or "2nd"
            
            # Process each course for this semester
            for course in courses_by_semester[semester]:
                # Populate the table with data for this specific course and semester
                populate_callback(table_widget, course, semester)
                
                # Create sheet even if no data (faculty can add data later)
                stats = get_course_statistics(user_id, course, semester, populate_callback, table_widget)
                course_statistics[f"{semester}_{course}"] = stats
                
                # Create sheet name: "1st_MATH101" or "2nd_ENG201"
                safe_course_name = course[:25]  # Limit course name length
                sheet_name = f"{semester_prefix}_{safe_course_name}"
                
                sheet = workbook.create_sheet(title=sheet_name)
                created_sheets.append(sheet_name)
                
                # Add title with consistent styling
                sheet.row_dimensions[1].height = 40
                sheet.cell(row=1, column=1, value=f"Subject: {course} - {semester}")
                sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(9, table_widget.columnCount() + 3))
                title_cell = sheet.cell(row=1, column=1)
                title_cell.font = title_font
                title_cell.alignment = Alignment(horizontal="center", vertical="center")
                title_cell.fill = PatternFill(start_color="E5E5E5", end_color="E5E5E5", fill_type="solid")
                
                # Add a link back to the summary
                sheet.cell(row=2, column=1, value="← Back to Summary")
                cell = sheet.cell(row=2, column=1)
                cell.hyperlink = "#Summary!A1"
                cell.font = Font(color="0563C1", underline="single")
                
                # Add Enhanced Faculty Grading System Configuration Area
                sheet.cell(row=3, column=1, value="SMART GWA CALCULATION SYSTEM")
                sheet.cell(row=3, column=1).font = Font(bold=True, size=14, color="4CAF50")
                sheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=6)
                
                # Explanation of smart GWA calculation
                sheet.cell(row=4, column=1, value="• Automatically adjusts weights when scores are missing")
                sheet.cell(row=4, column=1).font = Font(size=10, italic=True)
                sheet.merge_cells(start_row=4, start_column=1, end_row=4, end_column=6)
                
                sheet.cell(row=5, column=1, value="• Uses proportional weighting based on available exam scores")
                sheet.cell(row=5, column=1).font = Font(size=10, italic=True)
                sheet.merge_cells(start_row=5, start_column=1, end_row=5, end_column=6)
                
                # Grading system configuration
                sheet.cell(row=6, column=1, value="Passing Grade:")
                sheet.cell(row=6, column=1).font = Font(bold=True)
                sheet.cell(row=6, column=2, value=75)  # Default passing grade - faculty can modify
                sheet.cell(row=6, column=2).fill = grading_fill
                sheet.cell(row=6, column=2).border = thin_border
                
                sheet.cell(row=6, column=3, value="Standard Weights:")
                sheet.cell(row=6, column=3).font = Font(bold=True)
                sheet.cell(row=6, column=4, value="Prelim 30%")
                sheet.cell(row=6, column=4).fill = grading_fill
                sheet.cell(row=6, column=4).border = thin_border
                
                sheet.cell(row=6, column=5, value="Midterm 30%")
                sheet.cell(row=6, column=5).fill = grading_fill
                sheet.cell(row=6, column=5).border = thin_border
                
                sheet.cell(row=6, column=6, value="Finals 40%")
                sheet.cell(row=6, column=6).fill = grading_fill
                sheet.cell(row=6, column=6).border = thin_border
                
                # Add current enhanced statistics
                stats_text = f"Total: {stats['total_students']} | Pass: {stats['passing_students']} | Fail: {stats['failing_students']} | Incomplete: {stats['incomplete_students']}"
                sheet.cell(row=7, column=1, value=stats_text)
                sheet.cell(row=7, column=1).font = Font(italic=True, size=10)
                sheet.merge_cells(start_row=7, start_column=1, end_row=7, end_column=6)
                
                # Find exam columns for GWA calculation
                prelim_col = midterm_col = finals_col = -1
                
                # Add headers (row 9) - include original headers plus GWA, Status, and Calculation Method
                header_row = 9
                if table_widget.columnCount() > 0:
                    for col in range(table_widget.columnCount()):
                        header_item = table_widget.horizontalHeaderItem(col)
                        header_text = header_item.text() if header_item else f"Column {col+1}"
                        
                        # Track exam columns
                        if "PRELIM" in header_text.upper():
                            prelim_col = col
                        elif "MIDTERM" in header_text.upper():
                            midterm_col = col
                        elif "FINAL" in header_text.upper():
                            finals_col = col
                        
                        cell = sheet.cell(row=header_row, column=col + 1, value=header_text)
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = header_alignment
                        cell.border = thin_border
                else:
                    # Create default headers if no data
                    default_headers = ["Full Name", "Prelim", "Midterm", "Finals"]
                    for col, header in enumerate(default_headers):
                        cell = sheet.cell(row=header_row, column=col + 1, value=header)
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = header_alignment
                        cell.border = thin_border
                        
                        # Set exam column references for default headers
                        if header == "Prelim":
                            prelim_col = col
                        elif header == "Midterm":
                            midterm_col = col
                        elif header == "Finals":
                            finals_col = col
                
                # Add enhanced GWA headers
                base_cols = max(table_widget.columnCount(), 4)
                gwa_col = base_cols + 1
                status_col = gwa_col + 1
                method_col = status_col + 1
                
                # GWA header
                gwa_header = sheet.cell(row=header_row, column=gwa_col, value="GWA")
                gwa_header.fill = gwa_header_fill
                gwa_header.font = Font(color="FFFFFF", bold=True, size=12)
                gwa_header.alignment = header_alignment
                gwa_header.border = thin_border
                
                # Status header
                status_header = sheet.cell(row=header_row, column=status_col, value="Status")
                status_header.fill = gwa_header_fill
                status_header.font = Font(color="FFFFFF", bold=True, size=12)
                status_header.alignment = header_alignment
                status_header.border = thin_border
                
                # Calculation Method header
                method_header = sheet.cell(row=header_row, column=method_col, value="Calculation Method")
                method_header.fill = gwa_header_fill
                method_header.font = Font(color="FFFFFF", bold=True, size=11)
                method_header.alignment = header_alignment
                method_header.border = thin_border
                
                # Add data (starting at row 10) with enhanced GWA calculation
                data_start_row = header_row + 1
                
                if table_widget.rowCount() > 0:
                    # Add existing data with enhanced calculations
                    for row in range(table_widget.rowCount()):
                        prelim_score = ""
                        midterm_score = ""
                        finals_score = ""
                        
                        # Add original data
                        for col in range(table_widget.columnCount()):
                            item = table_widget.item(row, col)
                            value = item.text() if item else ""
                            
                            # Store exam scores for GWA calculation
                            if col == prelim_col:
                                prelim_score = value
                            elif col == midterm_col:
                                midterm_score = value
                            elif col == finals_col:
                                finals_score = value
                            
                            cell = sheet.cell(row=data_start_row + row, column=col + 1, value=value)
                            cell.alignment = cell_alignment
                            cell.border = thin_border
                        
                        # Calculate enhanced GWA
                        gwa = calculate_gwa_with_available_scores(prelim_score, midterm_score, finals_score)
                        details = get_gwa_calculation_details(prelim_score, midterm_score, finals_score)
                        
                        # Add GWA (show value or "N/A" if no valid scores)
                        gwa_display = gwa if gwa is not None else "N/A"
                        gwa_cell = sheet.cell(row=data_start_row + row, column=gwa_col, value=gwa_display)
                        gwa_cell.alignment = cell_alignment
                        gwa_cell.border = thin_border
                        gwa_cell.font = Font(bold=True)
                        
                        # Add Status with enhanced color coding
                        status_text = get_grade_status_text(gwa)
                        status_cell = sheet.cell(row=data_start_row + row, column=status_col, value=status_text)
                        status_cell.alignment = cell_alignment
                        status_cell.border = thin_border
                        
                        # Apply enhanced styling based on status
                        if status_text == "PASS":
                            status_cell.fill = pass_fill
                            status_cell.font = pass_font
                            gwa_cell.fill = pass_fill
                        elif status_text == "FAIL":
                            status_cell.fill = fail_fill
                            status_cell.font = fail_font
                            gwa_cell.fill = fail_fill
                        else:  # INCOMPLETE
                            status_cell.fill = incomplete_fill
                            status_cell.font = incomplete_font
                            gwa_cell.fill = incomplete_fill
                        
                        # Add calculation method
                        method_cell = sheet.cell(row=data_start_row + row, column=method_col, value=details['calculation_method'])
                        method_cell.alignment = cell_alignment
                        method_cell.border = thin_border
                        method_cell.font = Font(size=9, italic=True)
                else:
                    # Add empty rows for faculty to fill in
                    for row in range(10):  # Add 10 empty rows
                        for col in range(method_col):
                            cell = sheet.cell(row=data_start_row + row, column=col + 1, value="")
                            cell.alignment = cell_alignment
                            cell.border = thin_border
                
                # Auto-size columns including new enhanced columns
                for col in range(1, method_col + 1):
                    max_length = 0
                    for row in range(1, data_start_row + max(table_widget.rowCount(), 10) + 1):
                        cell_value = sheet.cell(row=row, column=col).value
                        if cell_value:
                            max_length = max(max_length, len(str(cell_value)))
                    # Set minimum width for calculation method column
                    min_width = 25 if col == method_col else 12
                    sheet.column_dimensions[get_column_letter(col)].width = max(max_length + 2, min_width)
        
        # Enhanced Summary sheet
        current_row = 4
        
        # Add Performance Overview section
        summary_sheet.cell(row=current_row, column=1, value="Enhanced Course Performance Overview")
        summary_sheet.cell(row=current_row, column=1).font = subtitle_font
        summary_sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
        current_row += 1
        
        # Add enhanced headers
        headers = ["Course", "Semester", "Pass", "Fail", "Incomplete", "Status", "Action"]
        for col, header in enumerate(headers, 1):
            cell = summary_sheet.cell(row=current_row, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
        current_row += 1
        
        # Add data for each course in both semesters
        for semester in ["1st Semester", "2nd Semester"]:
            semester_prefix = semester[:3]
            
            # Add semester separator row
            summary_sheet.cell(row=current_row, column=1, value=f"{semester} Courses")
            summary_sheet.cell(row=current_row, column=1).font = Font(bold=True, color="0066CC")
            summary_sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(headers))
            summary_sheet.cell(row=current_row, column=1).fill = PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid")
            current_row += 1
            
            # Add courses for this semester
            semester_courses = courses_by_semester[semester]
            if not semester_courses:
                # Show "No courses" message
                summary_sheet.cell(row=current_row, column=1, value="No courses available")
                summary_sheet.cell(row=current_row, column=1).font = Font(italic=True, color="888888")
                summary_sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(headers))
                for col in range(1, len(headers) + 1):
                    summary_sheet.cell(row=current_row, column=col).border = thin_border
                current_row += 1
            else:
                for course in semester_courses:
                    sheet_name = f"{semester_prefix}_{course}"
                    stats_key = f"{semester}_{course}"
                    
                    # Course name
                    summary_sheet.cell(row=current_row, column=1, value=course)
                    summary_sheet.cell(row=current_row, column=1).border = thin_border
                    summary_sheet.cell(row=current_row, column=1).alignment = cell_alignment
                    
                    # Semester
                    summary_sheet.cell(row=current_row, column=2, value=semester)
                    summary_sheet.cell(row=current_row, column=2).border = thin_border
                    summary_sheet.cell(row=current_row, column=2).alignment = cell_alignment
                    
                    # Enhanced statistics display
                    if stats_key in course_statistics:
                        stats = course_statistics[stats_key]
                        
                        # Pass count
                        summary_sheet.cell(row=current_row, column=3, value=stats['passing_students'])
                        summary_sheet.cell(row=current_row, column=3).border = thin_border
                        summary_sheet.cell(row=current_row, column=3).alignment = cell_alignment
                        summary_sheet.cell(row=current_row, column=3).fill = pass_fill
                        
                        # Fail count
                        summary_sheet.cell(row=current_row, column=4, value=stats['failing_students'])
                        summary_sheet.cell(row=current_row, column=4).border = thin_border
                        summary_sheet.cell(row=current_row, column=4).alignment = cell_alignment
                        summary_sheet.cell(row=current_row, column=4).fill = fail_fill
                        
                        # Incomplete count
                        summary_sheet.cell(row=current_row, column=5, value=stats['incomplete_students'])
                        summary_sheet.cell(row=current_row, column=5).border = thin_border
                        summary_sheet.cell(row=current_row, column=5).alignment = cell_alignment
                        summary_sheet.cell(row=current_row, column=5).fill = incomplete_fill
                        
                        # Enhanced status based on comprehensive analysis
                        total_with_scores = stats['passing_students'] + stats['failing_students']
                        if total_with_scores == 0:
                            status = "No Data Available"
                            status_color = "808080"  # Gray
                        elif stats['incomplete_students'] > total_with_scores:
                            status = "Mostly Incomplete"
                            status_color = "FF8C00"  # Dark orange
                        elif stats['passing_students'] > stats['failing_students']:
                            if stats['incomplete_students'] == 0:
                                status = "Excellent Performance"
                                status_color = "006400"  # Dark green
                            else:
                                status = "Good Performance"
                                status_color = "228B22"  # Forest green
                        elif stats['failing_students'] > stats['passing_students']:
                            status = "Needs Attention"
                            status_color = "DC143C"  # Crimson
                        else:
                            status = "Mixed Performance"
                            status_color = "FF8C00"  # Dark orange
                        
                        status_cell = summary_sheet.cell(row=current_row, column=6, value=status)
                        status_cell.font = Font(color=status_color, bold=True)
                        status_cell.border = thin_border
                        status_cell.alignment = cell_alignment
                        
                        # Action link
                        action_cell = summary_sheet.cell(row=current_row, column=7, value="Open Sheet")
                        action_cell.hyperlink = f"#{sheet_name}!A1"
                        action_cell.font = Font(color="0563C1", underline="single")
                        action_cell.border = thin_border
                        action_cell.alignment = cell_alignment
                    
                    current_row += 1
            
            # Add spacing between semesters
            current_row += 1
        
        # Add legend/explanation section
        current_row += 1
        summary_sheet.cell(row=current_row, column=1, value="SMART GWA CALCULATION LEGEND")
        summary_sheet.cell(row=current_row, column=1).font = Font(bold=True, size=12, color="4CAF50")
        summary_sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
        current_row += 1
        
        legend_items = [
            "• All 3 scores available: Uses standard weights (Prelim 30%, Midterm 30%, Finals 40%)",
            "• 2 scores available: Redistributes weights proportionally",
            "• 1 score available: Uses that score as 100% of the grade",
            "• No scores available: Shows 'INCOMPLETE' status",
            "• PASS: GWA ≥ 75 | FAIL: GWA < 75 | INCOMPLETE: No valid scores"
        ]
        
        for item in legend_items:
            summary_sheet.cell(row=current_row, column=1, value=item)
            summary_sheet.cell(row=current_row, column=1).font = Font(size=10)
            summary_sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
            current_row += 1
        
        # Auto-size columns in summary sheet
        for col in range(1, len(headers) + 1):
            max_length = 0
            for row in range(1, current_row + 1):
                cell_value = summary_sheet.cell(row=row, column=col).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            summary_sheet.column_dimensions[get_column_letter(col)].width = max(max_length + 4, 12)

        # Save the workbook to the designated path
        file_path = get_excel_path(user_id)
        workbook.save(file_path)
        del workbook  
        print(f"Enhanced Excel file with Smart GWA calculations saved successfully at {file_path}")
        return file_path

    except Exception as e:
        print(f"An error occurred while saving the Enhanced Excel file: {e}")
        raise e